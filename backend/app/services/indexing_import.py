from __future__ import annotations

import csv
import io
import re
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook

URL_HEADER_NAMES = {
    "url", "urls", "page", "pages", "address", "link", "links",
    "страница", "страницы", "адрес", "адрес страницы", "url страницы",
    "url-адрес", "url адрес", "ссылка", "ссылки", "landing page",
}


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def _normalize_header(value: Any) -> str:
    return _clean(value).lower().replace("ё", "е")


def _looks_like_url(value: str) -> bool:
    value = value.strip()
    if not value or " " in value:
        return False
    try:
        parsed = urlparse(value)
    except Exception:
        return False
    return parsed.scheme.lower() in {"http", "https"} and bool(parsed.netloc)


def _host(value: str) -> str:
    return urlparse(value).netloc.lower().removeprefix("www.")


def _detect_header_row(rows: list[list[str]]) -> int:
    # Prefer a row that contains a known GSC-style URL header. If nothing
    # resembles a header, the safest fallback is the first non-empty row;
    # guessing from values can accidentally promote a data row to headers.
    for index, row in enumerate(rows[:12]):
        normalized = [_normalize_header(cell) for cell in row]
        if any(cell in URL_HEADER_NAMES for cell in normalized):
            return index
        if any(cell == "url" or cell.startswith("url ") or cell.startswith("url-") for cell in normalized):
            return index
        if any(cell in {"страница", "page", "address", "адрес"} for cell in normalized):
            return index
    return 0


def _unique_headers(raw_headers: list[str]) -> list[str]:
    result: list[str] = []
    counts: dict[str, int] = {}
    for index, raw in enumerate(raw_headers, start=1):
        base = _clean(raw) or f"Колонка {index}"
        key = base.lower()
        counts[key] = counts.get(key, 0) + 1
        result.append(base if counts[key] == 1 else f"{base} ({counts[key]})")
    return result


def _detect_url_column(headers: list[str], rows: list[list[str]]) -> str | None:
    # Prefer common GSC names first.
    for header in headers:
        normalized = _normalize_header(header)
        if normalized in URL_HEADER_NAMES:
            return header
    for header in headers:
        normalized = _normalize_header(header)
        if "url" in normalized or normalized in {"страница", "page", "address", "адрес"}:
            return header

    # Otherwise use value heuristics. Require a reasonably strong URL ratio.
    best: tuple[float, str] | None = None
    for col_index, header in enumerate(headers):
        values = [_clean(row[col_index]) for row in rows[:200] if col_index < len(row) and _clean(row[col_index])]
        if not values:
            continue
        url_count = sum(1 for value in values if _looks_like_url(value))
        ratio = url_count / len(values)
        if url_count >= 2 and ratio >= 0.6 and (best is None or ratio > best[0]):
            best = (ratio, header)
    return best[1] if best else None


def _decode_csv(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _read_csv(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = _decode_csv(content)
    sample = text[:8000]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = None
    reader = csv.reader(io.StringIO(text), dialect) if dialect is not None else csv.reader(io.StringIO(text), delimiter=";")
    rows = [[_clean(cell) for cell in row] for row in reader]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        raise ValueError("Файл CSV пуст.")
    header_index = _detect_header_row(rows)
    raw_headers = rows[header_index]
    headers = _unique_headers(raw_headers)
    data = rows[header_index + 1 :]
    width = len(headers)
    data = [(row + [""] * width)[:width] for row in data]
    return headers, data


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    fallback: tuple[list[str], list[list[str]]] | None = None

    # GSC exports can contain more than one sheet. Prefer the visible sheet that
    # actually contains a detectable URL column instead of blindly taking the
    # first worksheet (which may contain metadata or a chart table).
    for sheet in workbook.worksheets:
        if sheet.sheet_state != "visible" or not sheet.max_row:
            continue
        raw_rows: list[list[str]] = []
        max_col = min(max(int(sheet.max_column or 1), 1), 100)
        for row in sheet.iter_rows(values_only=True, max_col=max_col):
            values = [_clean(cell) for cell in row]
            if any(values):
                raw_rows.append(values)
            if len(raw_rows) >= 10000:
                break
        if not raw_rows:
            continue
        header_index = _detect_header_row(raw_rows)
        headers = _unique_headers(raw_rows[header_index])
        width = len(headers)
        data = [(row + [""] * width)[:width] for row in raw_rows[header_index + 1 :]]
        parsed = (headers, data)
        if fallback is None:
            fallback = parsed
        if _detect_url_column(headers, data):
            return parsed

    if fallback is not None:
        return fallback
    raise ValueError("В XLSX не найден лист с данными.")


def parse_gsc_file(content: bytes, filename: str, project_domain: str, url_column: str = "") -> dict[str, Any]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        headers, rows = _read_csv(content)
    elif lower.endswith(".xlsx"):
        headers, rows = _read_xlsx(content)
    else:
        raise ValueError("Поддерживаются только XLSX и CSV.")

    detected = _detect_url_column(headers, rows)
    selected = url_column.strip() or detected or ""
    if selected and selected not in headers:
        raise ValueError("Выбранная колонка не найдена в файле.")

    preview = []
    for row in rows[:8]:
        preview.append({headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))})

    base_result: dict[str, Any] = {
        "columns": headers,
        "detected_column": detected,
        "selected_column": selected,
        "needs_column": not bool(selected),
        "preview": preview,
        "rows_total": len(rows),
    }
    if not selected:
        return {
            **base_result,
            "found_urls": 0,
            "unique_urls": 0,
            "duplicates": 0,
            "invalid_urls": 0,
            "other_domain_urls": 0,
            "urls": [],
            "invalid_values": [],
            "other_domain_values": [],
        }

    index = headers.index(selected)
    raw_values = [_clean(row[index]) for row in rows if index < len(row) and _clean(row[index])]
    project_host = _host(project_domain)
    valid_same_domain: list[str] = []
    invalid_values: list[str] = []
    other_domain_values: list[str] = []
    seen: set[str] = set()
    duplicate_count = 0

    for value in raw_values:
        if not _looks_like_url(value):
            invalid_values.append(value)
            continue
        if _host(value) != project_host:
            other_domain_values.append(value)
            continue
        # Preserve the original URL spelling for the HTTP check, but dedupe exact
        # normalized scheme/host/path/query values case-insensitively on host only.
        parsed = urlparse(value)
        normalized = parsed._replace(scheme=parsed.scheme.lower(), netloc=parsed.netloc.lower(), fragment="").geturl()
        key = normalized
        if key in seen:
            duplicate_count += 1
            continue
        seen.add(key)
        valid_same_domain.append(normalized)

    return {
        **base_result,
        "found_urls": len(raw_values),
        "unique_urls": len(valid_same_domain),
        "duplicates": duplicate_count,
        "invalid_urls": len(invalid_values),
        "other_domain_urls": len(other_domain_values),
        "urls": valid_same_domain,
        "invalid_values": invalid_values[:100],
        "other_domain_values": other_domain_values[:100],
    }
