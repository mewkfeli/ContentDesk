from __future__ import annotations

import io
import json
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

from app.db.database import get_connection
from app.services.indexing_import import parse_gsc_file
from app.services.meta_description_audit import preview_sitemap, suggest_description_details

router = APIRouter(prefix="/meta-description-audits", tags=["Meta Description Audit"])


def _project(project_id: int):
    with get_connection() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
        settings = conn.execute("SELECT settings_json FROM app_settings WHERE id=1").fetchone()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    cfg = json.loads(settings["settings_json"]) if settings else {}
    patterns = [x.strip() for x in (cfg.get("global_excludes", "") + "\n" + (project["exclude_patterns"] or "")).splitlines() if x.strip()]
    return project, patterns


class SitemapPreviewRequest(BaseModel):
    project_id: int
    sitemap_url: str = Field(default="", max_length=2048)
    max_urls: int = Field(default=5000, ge=1, le=20000)


@router.post("/preview-sitemap")
async def sitemap_preview(payload: SitemapPreviewRequest):
    project, patterns = _project(payload.project_id)
    return await preview_sitemap(project["domain"], payload.sitemap_url or project["sitemap_url"] or "", payload.max_urls, patterns)


@router.post("/import")
async def import_urls(project_id: int = Form(...), file: UploadFile = File(...), url_column: str = Form("")):
    project, _ = _project(project_id)
    if not file.filename: raise HTTPException(status_code=400, detail="Выберите XLSX или CSV файл")
    content = await file.read()
    try:
        result = parse_gsc_file(content, file.filename, project["domain"], url_column)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"project_id": project_id, "project_name": project["name"], "project_domain": project["domain"], "filename": file.filename, **result}


@router.get("")
def list_reports(project_id: int | None = None, limit: int = 30):
    where = "WHERE a.project_id=?" if project_id else ""
    params: tuple[Any, ...] = (project_id, limit) if project_id else (limit,)
    with get_connection() as conn:
        rows = conn.execute(f"""SELECT a.*,p.name project_name,p.domain project_domain FROM meta_description_audits a JOIN projects p ON p.id=a.project_id {where} ORDER BY a.id DESC LIMIT ?""", params).fetchall()
    return [dict(row) | {"result_json": None} for row in rows]


@router.get("/{report_id}")
def get_report(report_id: int):
    with get_connection() as conn:
        row = conn.execute("SELECT a.*,p.name project_name,p.domain project_domain FROM meta_description_audits a JOIN projects p ON p.id=a.project_id WHERE a.id=?", (report_id,)).fetchone()
    if not row: raise HTTPException(status_code=404, detail="Отчёт не найден")
    result = json.loads(row["result_json"])
    base = dict(row); base.pop("result_json", None)
    return {**base, **result}


class GenerateRequest(BaseModel):
    urls: list[str] = Field(min_length=1, max_length=2000)


@router.post("/{report_id}/generate")
def generate(report_id: int, payload: GenerateRequest):
    with get_connection() as conn:
        row = conn.execute("SELECT result_json FROM meta_description_audits WHERE id=?", (report_id,)).fetchone()
        if not row: raise HTTPException(status_code=404, detail="Отчёт не найден")
        result = json.loads(row["result_json"])
        wanted = set(payload.urls)
        generated = []
        for item in result.get("rows", []):
            if item.get("url") not in wanted: continue
            details = suggest_description_details(item)
            suggestion = details.get("description", "")
            item["suggested_description"] = suggestion
            item["generation_used_facts"] = details.get("used_facts", [])
            item["generation_notes"] = details.get("notes", [])
            item["suggestion_action"] = "review" if suggestion else ""
            item["generation_blocked_reason"] = "" if suggestion else (details.get("notes", [""])[0] if details.get("notes") else "Недостаточно данных страницы")
            generated.append({"url": item["url"], "suggested_description": suggestion, "used_facts": item["generation_used_facts"], "notes": item["generation_notes"], "blocked_reason": item.get("generation_blocked_reason", "")})
        conn.execute("UPDATE meta_description_audits SET result_json=? WHERE id=?", (json.dumps(result, ensure_ascii=False), report_id))
        conn.commit()
    return generated


class SuggestionUpdate(BaseModel):
    url: str
    suggested_description: str = ""
    action: str = Field(default="review", pattern="^(review|accepted|rejected)$")


@router.patch("/{report_id}/suggestion")
def update_suggestion(report_id: int, payload: SuggestionUpdate):
    with get_connection() as conn:
        row = conn.execute("SELECT result_json FROM meta_description_audits WHERE id=?", (report_id,)).fetchone()
        if not row: raise HTTPException(status_code=404, detail="Отчёт не найден")
        result = json.loads(row["result_json"])
        item = next((x for x in result.get("rows", []) if x.get("url") == payload.url), None)
        if not item: raise HTTPException(status_code=404, detail="URL не найден")
        item["suggested_description"] = payload.suggested_description.strip()
        item["suggestion_action"] = payload.action
        conn.execute("UPDATE meta_description_audits SET result_json=? WHERE id=?", (json.dumps(result, ensure_ascii=False), report_id))
        conn.commit()
    return item


@router.get("/{report_id}/export.xlsx")
def export_xlsx(report_id: int, status: str = "", issue: str = "", section: str = "", search: str = "", page_type: str = "", indexable: str = ""):
    with get_connection() as conn:
        db = conn.execute("SELECT a.*,p.name project_name FROM meta_description_audits a JOIN projects p ON p.id=a.project_id WHERE a.id=?", (report_id,)).fetchone()
    if not db: raise HTTPException(status_code=404, detail="Отчёт не найден")
    result = json.loads(db["result_json"])
    rows = result.get("rows", [])
    if status: rows = [r for r in rows if r.get("status") == status]
    if issue: rows = [r for r in rows if issue in r.get("issues", [])]
    if section: rows = [r for r in rows if r.get("section") == section]
    if search: rows = [r for r in rows if search.lower() in r.get("url", "").lower()]
    if page_type: rows = [r for r in rows if r.get("page_type", "unknown") == page_type]
    if indexable: rows = [r for r in rows if r.get("indexable", "unknown") == indexable]

    wb = Workbook(); ws = wb.active; ws.title = "Аудит Description"
    headers = ["URL","Тип страницы","Индексируемая","HTTP-статус","Title","H1","Текущий Description","Description в HTML","Длина","Проблемы","Статус","Новый Description","Использованные данные","Примечания генерации","Действие"]
    ws.append(headers)
    fills = {"ok":"E9F7EF","review":"FFF4D6","replace":"FDE8E7","broken":"FADBD8","technical":"E5E7EB","template":"F3E8FF"}
    action_labels = {"ok":"Всё в порядке","review":"Проверить","replace":"Исправить","broken":"HTTP-ошибка / битая страница","technical":"Техническая проблема","template":"Проблема шаблона Description"}
    for r in rows:
        ws.append([
            r.get("url",""), r.get("page_type_label","Не определено"), r.get("indexable_label","Не определено"),
            r.get("status_code",0), r.get("title",""), r.get("h1",""), r.get("description",""), r.get("description_raw",""),
            r.get("description_length",0), " | ".join(r.get("issue_labels",[])), r.get("status_label",""),
            r.get("suggested_description",""),
            " | ".join(f"{x.get('label','')}: {x.get('value','')}" for x in r.get("generation_used_facts",[]) if isinstance(x,dict)),
            " | ".join(r.get("generation_notes",[])), action_labels.get(r.get("status"),"")
        ])
        fill=PatternFill("solid",fgColor=fills.get(r.get("status"),"FFFFFF"))
        for c in ws[ws.max_row]: c.fill=fill; c.alignment=Alignment(vertical="top",wrap_text=True)
    for c in ws[1]: c.fill=PatternFill("solid",fgColor="1F2937"); c.font=Font(color="FFFFFF",bold=True)
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
    for i,w in enumerate([48,22,18,14,38,38,65,65,10,48,24,65,55,45,24],1): ws.column_dimensions[get_column_letter(i)].width=w

    info=wb.create_sheet("Сводка")
    info.append(["Параметр","Значение"]); info.append(["Проект",db["project_name"]]); info.append(["Проверено",result.get("urls_total",0)])
    for key,label in [("ok","Без ошибок"),("review","Требуют проверки"),("replace","Требуют исправления"),("template","Проблема шаблона Description"),("broken","HTTP-ошибки / битые страницы"),("technical","Технические проблемы")]: info.append([label,result.get("status_counts",{}).get(key,0)])
    for key,label in [("missing","Description отсутствует"),("too_long","Слишком длинные"),("too_short","Слишком короткие"),("duplicate","Дубликаты"),("html_entities","HTML-сущности"),("emoji","Эмодзи / спецсимволы"),("template","Шаблонные")]: info.append([label,result.get("issue_counts",{}).get(key,0)])
    info.append(["Технических URL исключено", result.get("technical_excluded",0)])
    info.append(["Ошибок получения страниц", result.get("fetch_errors",0)])
    info.append(["URL с HTTP-ошибкой", result.get("http_errors",0)])
    info.append(["HTTP 404", result.get("http_404",0)])
    info.append(["HTTP 5xx", result.get("http_5xx",0)])
    info.append(["Страниц с проблемой шаблона", result.get("template_problem_count",0)])
    info.append(["Товаров для ручного исправления", result.get("products_content_fix",result.get("products_to_fix",0))])
    info.append(["Товаров с проблемой шаблона", result.get("products_template_problem",0)])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f'attachment; filename="meta-description-audit-{report_id}.xlsx"'})
