from __future__ import annotations

import io
import json
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.database import get_connection
from app.services.indexing_import import parse_gsc_file
from app.services.indexing_check import _classify

router = APIRouter(prefix="/indexing-checks", tags=["Indexing diagnostics"])


def _summary(row: Any) -> dict[str, Any]:
    return {
        "id": row["id"], "project_id": row["project_id"], "project_name": row["project_name"],
        "project_domain": row["project_domain"], "source_name": row["source_name"],
        "sitemap_url": row["sitemap_url"], "urls_total": row["urls_total"],
        "ok_count": row["ok_count"], "content_count": row["content_count"],
        "developer_count": row["developer_count"], "insufficient_count": row["insufficient_count"] if "insufficient_count" in row.keys() else 0,
        "created_at": row["created_at"],
    }


@router.post("/import")
async def import_gsc_file(
    project_id: int = Form(...),
    file: UploadFile = File(...),
    url_column: str = Form(""),
):
    with get_connection() as conn:
        project = conn.execute("SELECT * FROM projects WHERE id=?", (project_id,)).fetchone()
    if not project:
        raise HTTPException(status_code=404, detail="Проект не найден")
    if not file.filename:
        raise HTTPException(status_code=400, detail="Выберите XLSX или CSV файл")
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Файл слишком большой. Максимум 20 МБ.")
    try:
        result = parse_gsc_file(content, file.filename, project["domain"], url_column)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"project_id": project_id, "project_name": project["name"], "project_domain": project["domain"], "filename": file.filename, **result}


@router.get("")
def list_indexing_checks(project_id: int | None = None, limit: int = 30):
    limit = max(1, min(limit, 100))
    where = "WHERE c.project_id=?" if project_id is not None else ""
    params: tuple[Any, ...] = (project_id, limit) if project_id is not None else (limit,)
    with get_connection() as conn:
        rows = conn.execute(
            f"""
            SELECT c.*, p.name project_name, p.domain project_domain
            FROM indexing_checks c JOIN projects p ON p.id=c.project_id
            {where} ORDER BY c.id DESC LIMIT ?
            """, params,
        ).fetchall()
    return [_summary(row) for row in rows]


@router.get("/{report_id}")
def get_indexing_check(report_id: int):
    with get_connection() as conn:
        row = conn.execute(
            """
            SELECT c.*, p.name project_name, p.domain project_domain
            FROM indexing_checks c JOIN projects p ON p.id=c.project_id WHERE c.id=?
            """, (report_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Отчёт проверки индексации не найден")
    result = json.loads(row["result_json"])
    return {**_summary(row), **result}


def _donors_text(row: dict[str, Any]) -> str:
    values = []
    for item in row.get("incoming_links", []):
        source = item.get("source", "")
        anchor = item.get("anchor", "")
        link_type = item.get("type", "other")
        values.append(f"{source} — {anchor} — {link_type}" if anchor else f"{source} — {link_type}")
    return "\n".join(values)


class HubConfirmation(BaseModel):
    url: str
    hub_url: str


@router.patch("/{report_id}/hub")
def confirm_indexing_hub(report_id: int, payload: HubConfirmation):
    with get_connection() as conn:
        row = conn.execute("SELECT result_json FROM indexing_checks WHERE id=?", (report_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Отчёт проверки индексации не найден")
        result = json.loads(row["result_json"])
        target = next((item for item in result.get("rows", []) if item.get("url") == payload.url), None)
        if not target:
            raise HTTPException(status_code=404, detail="URL не найден в отчёте")
        donor = next((item for item in target.get("incoming_links", []) if item.get("source") == payload.hub_url), None)
        target["hub_candidate"] = payload.hub_url
        target["hub_kind"] = "confirmed"
        target["hub_status"] = "yes" if donor else "no"
        target["hub_confirmed"] = True
        if donor:
            donor["type"] = "hub"
        refreshed = _classify(target)
        target.clear(); target.update(refreshed)
        # Recompute report counts because confirming a real hub can remove a weak-inlinks warning.
        counts = {"ok": 0, "content": 0, "developer": 0, "insufficient": 0}
        for item in result.get("rows", []):
            counts[item.get("status", "ok")] = counts.get(item.get("status", "ok"), 0) + 1
        result["status_counts"] = counts
        conn.execute("UPDATE indexing_checks SET ok_count=?, content_count=?, developer_count=?, insufficient_count=?, result_json=? WHERE id=?", (
            counts.get("ok",0), counts.get("content",0), counts.get("developer",0), counts.get("insufficient",0), json.dumps(result, ensure_ascii=False), report_id
        ))
        conn.commit()
    return target


@router.get("/{report_id}/export.xlsx")
def export_indexing_xlsx(report_id: int):
    with get_connection() as conn:
        db_row = conn.execute(
            """
            SELECT c.*, p.name project_name, p.domain project_domain
            FROM indexing_checks c JOIN projects p ON p.id=c.project_id WHERE c.id=?
            """, (report_id,),
        ).fetchone()
    if not db_row:
        raise HTTPException(status_code=404, detail="Отчёт проверки индексации не найден")
    result = json.loads(db_row["result_json"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Проверка индексации"
    headers = [
        "URL", "Итог", "Исполнитель", "HTTP", "Конечный URL", "Robots", "X-Robots-Tag",
        "Canonical", "Sitemap", "Sitemap URL", "priority", "changefreq", "Inlinks", "Самоссылка", "Анкоры самоссылки", "Глубина",
        "Причина глубины", "Найден при crawl", "Ссылка с Главной", "Хаб", "Title", "H1", "Слов", "Проблемы", "Страницы-доноры", "Рекомендация",
    ]
    ws.append(headers)
    fills = {"ok": "E9F7EF", "content": "FFF4D6", "developer": "FDE8E7", "insufficient": "F2F4F7"}
    for row in result.get("rows", []):
        sitemap = row.get("sitemap", {})
        ws.append([
            row.get("url", ""), row.get("status_label", ""), row.get("executor", ""),
            row.get("initial_status_code") or row.get("status_code", ""), row.get("final_url", ""),
            row.get("robots", ""), row.get("x_robots", ""), row.get("canonical", ""),
            "Да" if sitemap.get("present") else "Нет", sitemap.get("sitemap_url", ""),
            sitemap.get("priority", ""), sitemap.get("changefreq", ""), row.get("inlinks", 0),
            "Да" if row.get("self_link") else "Нет", " | ".join(row.get("self_link_anchors", [])),
            row.get("depth") if row.get("depth") is not None else "", row.get("depth_reason", ""),
            "Да" if row.get("found_in_crawl") else "Нет", "Да" if row.get("home_link") else "Нет",
            ("Не определён" if row.get("hub_status") == "unknown" else ("Да" if row.get("hub_status") == "yes" else "Нет")),
            row.get("title", ""), row.get("h1", ""), row.get("word_count", 0),
            " | ".join(row.get("problems", [])), _donors_text(row), row.get("recommendation", ""),
        ])
        fill = PatternFill("solid", fgColor=fills.get(row.get("status"), "FFFFFF"))
        for cell in ws[ws.max_row]:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [44, 30, 22, 10, 44, 22, 24, 44, 11, 44, 10, 12, 9, 12, 40, 9, 34, 15, 15, 14, 36, 36, 9, 46, 64, 64]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 32

    info = wb.create_sheet("Сводка")
    counts = result.get("status_counts", {})
    info.append(["Параметр", "Значение"])
    info.append(["Проект", db_row["project_name"]])
    info.append(["Домен", db_row["project_domain"]])
    info.append(["Источник", db_row["source_name"]])
    info.append(["URL проверено", result.get("urls_total", 0)])
    info.append(["Всё нормально", counts.get("ok", 0)])
    info.append(["Контент-менеджеру", counts.get("content", 0)])
    info.append(["Разработчику", counts.get("developer", 0)])
    info.append(["Недостаточно данных", counts.get("insufficient", 0)])
    crawl = result.get("crawl", {})
    info.append(["Crawl достаточен", "Да" if crawl.get("sufficient") else "Нет"])
    info.append(["Просканировано страниц", crawl.get("pages_crawled", crawl.get("pages_total", 0))])
    info.append(["Найдено HTML-ссылок", crawl.get("html_links_seen", crawl.get("links_total", 0))])
    info.append(["Уникальных URL найдено", crawl.get("unique_urls_found", 0)])
    info.append(["Ошибок обхода", crawl.get("errors_count", 0)])
    info.append(["Sitemap", result.get("sitemap_url", "")])
    for cell in info[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    info.column_dimensions["A"].width = 28
    info.column_dimensions["B"].width = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"indexing-check-{report_id}.xlsx"
    headers_out = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_out)
